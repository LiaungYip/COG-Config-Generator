import unittest
from typing import Mapping, Any, List

import openpyxl


def find_table(workbook: openpyxl.workbook.workbook.Workbook, table_name: str) \
        -> (openpyxl.worksheet.worksheet.Worksheet, openpyxl.worksheet.table.Table):
    # Given a Workbook and the name of that table, return the Worksheet the table is on, and the Table object.
    # Raises ValueError if the table doesn't exist.

    for worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
        for table in worksheet._tables:
            if table_name == table.name:
                return worksheet, table

    raise ValueError("Table %s doesn't exist." % table_name)

def get_table_data(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                   table: openpyxl.worksheet.table.Table) -> List[Mapping[str, Any]]:
    # Given a Worksheet, and a Table that is on that Worksheet,
    # Return the table data as a list of dictionaries.
    # Example: This Excel table -
    #
    # +---------+------------+--------+
    # | Colour  | Shape      | Number |
    # +---------+------------+--------+
    # | Red     | Square     | 1.2    |
    # | Yellow  | Triangle   | 3.4    |
    # | Blue    | Circle     | 5.6    |
    # +---------+------------+--------+
    #
    # becomes this list of dictionaries:
    # [
    #     {'Colour': 'Red', 'Shape': 'Square', 'Number': 1.2},
    #     {'Colour': 'Yellow', 'Shape': 'Triangle', 'Number': 3.4},
    #     {'Colour': 'Blue', 'Shape': 'Circle', 'Number': 5.6}
    # ]

    ref = table.ref  # i.e. ref = "C6:E9" - refers to entire table __including header row(s)__
    cell_array = worksheet[ref]  # cell_array = a nested list of cells
    # Example:
    # cell_array = (
    #   (<Cell 'APPLE'.C6>, <Cell 'APPLE'.D6>, <Cell 'APPLE'.E6>),
    #   (<Cell 'APPLE'.C7>, <Cell 'APPLE'.D7>, <Cell 'APPLE'.E7>),
    #   (<Cell 'APPLE'.C8>, <Cell 'APPLE'.D8>, <Cell 'APPLE'.E8>),
    #   (<Cell 'APPLE'.C9>, <Cell 'APPLE'.D9>, <Cell 'APPLE'.E9>)
    # )


    number_of_data_rows = len(cell_array) - 1  # not including header row

    if number_of_data_rows == 0:  # no data rows - I don't think this is actually possible for tables created in Microsoft Excel.
        return []

    data = [[cell.value for cell in row] for row in cell_array]

    data_rows = []
    header = data[0]
    for row in data[1:]:
        data_rows.append(dict(zip(header, row)))
    return data_rows


class TestExcelTables(unittest.TestCase):
    test_file_path = "./test_data/Test Workbook.xlsx"

    def test_load_worksheet(self):
        workbook = openpyxl.load_workbook(self.test_file_path)
        self.assertEqual(type(workbook), openpyxl.workbook.workbook.Workbook)

    def test_get_cell(self):
        workbook = openpyxl.load_workbook(self.test_file_path)
        worksheet = workbook["APPLE"]  # "APPLE" is the name of the worksheet. Note: Case sensitive.

        cell_A1 = worksheet["A1"]
        self.assertEqual(cell_A1.value, "This is cell A1.")

        cell_B2 = worksheet["B2"]
        self.assertEqual(cell_B2.value, "This is cell B2.")

    def test_no_such_table(self):
        workbook = openpyxl.load_workbook(self.test_file_path)
        with self.assertRaises(ValueError):
            find_table(workbook, "Nonexistent_Table_Name")

    def test_get_table(self):
        # There's a table called "Colour_Shape_And_Number" on the first worksheet of the test file.
        # It looks like this:
        # +---------+------------+--------+
        # | Colour  | Shape      | Number |
        # +---------+------------+--------+
        # | Red     | Square     | 1.2    |
        # | Yellow  | Triangle   | 3.4    |
        # | Blue    | Circle     | 5.6    |
        # +---------+------------+--------+
        expected = [
            {
                'Colour': 'Red',
                'Shape': 'Square',
                'Number': 1.2
            },
            {
                'Colour': 'Yellow',
                'Shape': 'Triangle',
                'Number': 3.4
            },
            {
                'Colour': 'Blue',
                'Shape': 'Circle',
                'Number': 5.6
            }
        ]
        workbook = openpyxl.load_workbook(self.test_file_path)
        worksheet, my_table = find_table(workbook, "Colour_Shape_And_Number")
        table_data = get_table_data(worksheet, my_table)

        self.assertEqual(table_data, expected)
        pass


if __name__ == '__main__':
    unittest.main()
